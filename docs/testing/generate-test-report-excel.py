#!/usr/bin/env python3
"""Xuat bao-cao-kiem-thu.xlsx (fallback khi PHP khong co ext-zip)."""
from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
PHP = Path(r"C:\xampp\php\php.exe")
DATA_PHP = ROOT / "docs" / "testing" / "test-report-data.php"
OUTPUT = ROOT / "docs" / "testing" / "bao-cao-kiem-thu.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(name="Times New Roman", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Times New Roman", size=11)
PASS_FILL = PatternFill("solid", fgColor="D4EDDA")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_data() -> dict:
    cmd = [str(PHP), "-r", f"echo json_encode(require '{DATA_PHP.as_posix()}', JSON_UNESCAPED_UNICODE);"]
    raw = subprocess.check_output(cmd, cwd=ROOT, text=True, encoding="utf-8")
    return json.loads(raw)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_table(ws, start_row: int, headers: list[str], rows: list[list]) -> int:
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if headers[c - 1] == "Kết quả" and val == "Pass":
                cell.fill = PASS_FILL
        r += 1
    return r


def autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    data = load_data()
    meta = data["meta"]
    sections = data["sections"]

    total = pass_count = fail_count = 0
    for section in sections:
        for case in section["cases"]:
            total += 1
            if case[4] == "Pass":
                pass_count += 1
            elif case[4] == "Fail":
                fail_count += 1
    rate = round(pass_count / total * 100, 1) if total else 0

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Tong quan"
    cover = [
        ["Bao cao kiem thu he thong", ""],
        [meta["project"], ""],
        [meta["subtitle"], ""],
        ["", ""],
        ["Phien ban", meta["version"]],
        ["Ngay kiem thu", meta["date"]],
        ["Nguoi thuc hien", meta["tester"]],
        ["Phuong phap", meta["method"]],
        ["", ""],
        ["Tong TC", "Pass", "Fail", "Ty le (%)"],
        [total, pass_count, fail_count, rate],
    ]
    for r, row in enumerate(cover, 1):
        for c, val in enumerate(row, 1):
            ws0.cell(row=r, column=c, value=val).font = BODY_FONT
    autosize(ws0, [28, 18, 12, 12])

    ws1 = wb.create_sheet("Tong hop module")
    mod_rows = []
    for i, s in enumerate(sections, 1):
        st = len(s["cases"])
        sp = sum(1 for c in s["cases"] if c[4] == "Pass")
        mod_rows.append([i, s["id"], s["title"], st, sp, st - sp, round(sp / st * 100, 1) if st else 0])
    mod_rows.append(["", "Tong", "", total, pass_count, fail_count, rate])
    write_table(ws1, 1, ["STT", "Module", "Ten module", "So TC", "Pass", "Fail", "Ty le (%)"], mod_rows)
    autosize(ws1, [6, 10, 42, 10, 10, 10, 12])

    ws2 = wb.create_sheet("Chi tiet test case")
    detail = []
    seq = 1
    for s in sections:
        for case in s["cases"]:
            detail.append([seq, case[0], s["id"], s["title"], case[1], case[2], case[3], case[4], case[5], "", "", ""])
            seq += 1
    write_table(
        ws2,
        1,
        ["STT", "Ma TC", "Module", "Ten module", "Chuc nang", "Cac buoc", "Ket qua mong doi", "Ket qua", "Uu tien", "Nguoi test", "Ngay test", "Ghi chu"],
        detail,
    )
    autosize(ws2, [5, 11, 8, 28, 22, 36, 38, 10, 12, 14, 12, 20])

    for s in sections:
        name = f"Module {s['id']}"[:31]
        ws = wb.create_sheet(name)
        rows = [[c[0], c[1], c[2], c[3], c[4], c[5]] for c in s["cases"]]
        write_table(ws, 1, ["Ma TC", "Chuc nang", "Cac buoc", "Ket qua mong doi", "Ket qua", "Uu tien"], rows)
        autosize(ws, [11, 24, 36, 40, 10, 12])

    wb.save(OUTPUT)
    print(str(OUTPUT))


if __name__ == "__main__":
    main()
